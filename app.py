import aiohttp
import asyncio
import re
import os
import requests
import math
import openpyxl
import pandas as pd
from datetime import datetime
from benedict import benedict
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

date = datetime.now().strftime("%d-%m")
URL = "https://api.hh.ru/vacancies"
headers = {
    "HH-User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36"}
params = {
    "text": "Python",
    "period": 7,
    "per_page": 100,
    "page": 0,
    "area": 113,
    "industry": 7,  # or 'professional_role': 11     # 'specialization': 1
    "experience": ["noExperience", "between1And3"],
    "schedule": "remote",
}


def remove_tags(text):
    """Remove HTML tags from a text string."""
    return re.sub(r"[<{].*?[}>]", "", text) if text else None


async def fetch_get(page_num_or_id, session):
    url = URL + \
        (f'/{page_num_or_id}' if isinstance(page_num_or_id, str) else '')
    var_params = {
        **params, 'page': page_num_or_id} if isinstance(page_num_or_id, int) else None
    async with session.get(url=url, headers=headers, params=var_params) as response:
        response.raise_for_status()
        data = await response.json()
        return data.get('items') if isinstance(page_num_or_id, int) else data


async def get_json(total_pages):
    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=20)
    ) as session:
        tasks = []
        if type(total_pages) == int:
            for page in range(total_pages):
                tasks.append(asyncio.create_task(fetch_get(page, session)))
            responses = await asyncio.gather(*tasks)
            return [item for sublist in responses for item in sublist]
        elif type(total_pages) == list:
            for id in total_pages:
                tasks.append(asyncio.create_task(fetch_get(id, session)))
            responses = await asyncio.gather(*tasks)
            return responses


def get_vacances():
    total_pages = calc_pages()
    if total_pages is None:
        return None
    data = asyncio.run(get_json(total_pages))
    new_arr = []
    for dic in data:
        dic = benedict(dic)        
        new_dict = {
            "id": dic["id"],
            "name": dic["name"],
            "area": dic["area.name"],
            "s_from": dic.get("salary.from"),
            "s_to": dic.get("salary.to"),
            "published_at": dic["published_at"],
            "employer": dic["employer.name"][:20],
            "requirement": remove_tags(dic["snippet.requirement"]),
            "responsibility": remove_tags(dic["snippet.responsibility"]),
        }
        new_arr.append(new_dict)

    df = pd.DataFrame(data=new_arr)
    df = df.drop_duplicates(subset=["name", "requirement"])
    df = df[df["name"].apply(lambda name: True if all([not re.search(word, name) for word in {"iddle", "имлид", "enior"}]) else False)]

    df["published_at"] = (
        pd.to_datetime(df["published_at"])
        .dt.tz_convert("Asia/Tokyo")
        .dt.tz_localize(None)
    )

    return df.set_index("id").sort_values(by="published_at", ascending=False)


def calc_pages():
    response = requests.get(URL, params=params, headers=headers, timeout=5)
    if not response.ok:
        print(response.status_code)
        return None

    answer = math.floor(response.json().get("found") / 100)
    return answer if 0 < answer < 20 else None


def enricher_key_skills(df):
    data = asyncio.run(get_json(df.index.tolist()))

    for vacancy in data:
        for key in list(vacancy.keys()):
            if key not in ["id", "key_skills", "description"]:
                del vacancy[key]
        vacancy["key_skills"] = ", ".join(
            [small_dict.get("name") for small_dict in vacancy["key_skills"] if small_dict])
        vacancy["description"] = remove_tags(vacancy["description"])
    new_df = df.merge(pd.DataFrame(data).set_index(
        "id"), how="left", left_index=True, right_index=True)
    new_df.insert(6, "key_skills", new_df.pop("key_skills"))
    return new_df


def format_worksheet(df):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df.reset_index(), index=False, header=True):
        ws.append(row)

    sheet = wb.active
    dict_col = {sheet[f"{openpyxl.utils.cell.get_column_letter(col)}1"].value: openpyxl.utils.cell.get_column_letter(col)
        for col in range(1, sheet.max_column + 1)
    }
    for name, value in dict_col.items():
        cell = f"{value}1"
        sheet[cell].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        column_width = {
            "name": 60,
            "area": 20,
            "published_at": 20,
            "key_skills": 90,
            "employer": 20,
            "requirement": 100,
            "responsibility": 100,
            "description": 100,
        }.get(name, 12)
        sheet.column_dimensions[value].width = column_width

    sheet.row_dimensions[1].height = 40

    tab = Table(displayName=f"Table", ref=f"A1:{openpyxl.utils.cell.get_column_letter(sheet.max_column)}{sheet.max_row}")
    style = TableStyleInfo(
        name="TableStyleLight10",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    # codiga-disable
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    os.makedirs("./files", exist_ok=True)
    wb.save(f'./files/{date}.xlsx')
    return print('All files saved')


if __name__ == "__main__":
    df = enricher_key_skills(get_vacances())
    print(df.shape[0], 'vacancies')
    format_worksheet(df)
